"""
RCC Mobile API — FastAPI backend for PWA mobile app.
Reads same RCC_DATA.xlsx and serves JSON endpoints.
Run: python api.py
"""

import math
import os
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

APP_DIR = Path(__file__).resolve().parent
DATA_FILE = APP_DIR / "RCC_DATA.xlsx"
SHEET_NAME = "MAIN"

# ── CLOUD MODE — download from OneDrive if ONEDRIVE_SHARE_URL is set ──
CLOUD_MODE = os.environ.get("RCC_CLOUD", "0") == "1"
ONEDRIVE_SHARE_URL = os.environ.get("ONEDRIVE_SHARE_URL", "")

def _sync_from_onedrive():
    """Download RCC_DATA.xlsx from OneDrive share link (cloud mode)."""
    import requests
    if not ONEDRIVE_SHARE_URL:
        return False
    try:
        share_url = ONEDRIVE_SHARE_URL.strip()
        
        # Append &download=1 to force direct download from OneDrive
        separator = "&" if "?" in share_url else "?"
        download_url = f"{share_url}{separator}download=1"
        
        print(f"☁️ Attempting OneDrive download...")
        resp = requests.get(download_url, timeout=60, allow_redirects=True)
        
        # Verify it's actual Excel content (not HTML error page)
        if resp.status_code == 200 and len(resp.content) > 1000:
            if b'<!DOCTYPE' in resp.content[:200] or b'<html' in resp.content[:200]:
                print(f"⚠️ Got HTML instead of Excel. Download failed.")
                return False
            
            DATA_FILE.write_bytes(resp.content)
            print(f"☁️ Fresh data downloaded from OneDrive ({len(resp.content)} bytes)")
            
            # Re-apply pending payment updates that haven't been synced to HDFC yet
            _reapply_pending_payments()
            
            return True
        else:
            print(f"⚠️ OneDrive download failed (HTTP {resp.status_code})")
            return False
    except Exception as e:
        print(f"⚠️ OneDrive sync error: {e}")
        return False

# ── CACHING — Excel sirf tab reload hoga jab file change ho ──
_cache = {"df": None, "mtime": 0}

def _reapply_pending_payments():
    """After OneDrive download, re-apply pending payments from Google Sheets to in-memory cache."""
    try:
        import requests as _req
        resp = _req.get(GSHEET_WEBHOOK, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            entries = data.get("entries", [])
            if not entries:
                return
            
            df = _cache.get("df")
            if df is None:
                return
            
            applied = 0
            for entry in entries:
                loan_no = str(entry.get("loan_no", "")).strip()
                mask = df["LOAN NO"].astype(str).str.strip() == loan_no
                if mask.any():
                    idx = df[mask].index[0]
                    if "Mode Of Payment" in df.columns:
                        df.at[idx, "Mode Of Payment"] = entry.get("mode", "")
                    if "Payment Date" in df.columns:
                        df.at[idx, "Payment Date"] = entry.get("date", "")
                    if "Paid Amount" in df.columns:
                        df.at[idx, "Paid Amount"] = float(entry.get("amount", 0))
                    if "RECEIPT CUT" in df.columns:
                        df.at[idx, "RECEIPT CUT"] = "PAID"
                    applied += 1
            
            if applied > 0:
                print(f"💳 Re-applied {applied} pending payments from GSheets")
    except Exception as e:
        print(f"⚠️ Re-apply payments error: {e}")

CREDENTIALS = {
    "ADMIN": {"password": "Admin@123", "role": "admin"},
    "AKSHAY KARALE": {"password": "akshay@123", "role": "executive"},
    "AMIT GADE": {"password": "amit@123", "role": "executive"},
    "AMOL DHUMAL": {"password": "amol@123", "role": "executive"},
    "ANWAR SHAIKH": {"password": "anwar@123", "role": "executive"},
    "HARIDAS DIVATE": {"password": "haridas@123", "role": "executive"},
    "HEMANT WALUNJ": {"password": "hemant@123", "role": "executive"},
    "KIRAN KHAIRNAR": {"password": "kiran@123", "role": "executive"},
    "NITIN KADAM": {"password": "nitin@123", "role": "executive"},
    "PARMESHWAR KOTULE": {"password": "parmeshwar@123", "role": "executive"},
    "SACHIN INGALE": {"password": "sachini@123", "role": "executive"},
    "SACHIN KHAPRE": {"password": "sachink@123", "role": "executive"},
    "SAGAR DONGRE": {"password": "sagar@123", "role": "executive"},
    "SANDEEP KHAIRNAR": {"password": "sandeep@123", "role": "executive"},
    "SUNNY SHINDE": {"password": "sunny@123", "role": "executive"},
    "SWAPNIL JADHAV": {"password": "swapnil@123", "role": "executive"},
    "TANAJI SURVASE": {"password": "tanaji@123", "role": "executive"},
    "UDAY SOHANE": {"password": "uday@123", "role": "executive"},
    "VIKRAM GAIKWAD": {"password": "vikram@123", "role": "executive"},
}

app = FastAPI(title="RCC Mobile API")

@app.on_event("startup")
async def startup_preload():
    """Pre-load Excel data when server starts (for Render/cloud)."""
    if CLOUD_MODE and ONEDRIVE_SHARE_URL:
        print("☁️ Cloud mode — downloading from OneDrive...")
        _sync_from_onedrive()
    load_data()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve PWA static files (no-cache during dev)
from fastapi.responses import Response
from starlette.middleware.base import BaseHTTPMiddleware

class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if "/mobile/" in str(request.url):
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        return response

app.add_middleware(NoCacheMiddleware)

# ── VISITOR TRACKING (file-based — persists across restarts) ──
from datetime import datetime as _dt, timedelta as _td
import json as _json

_IST_OFFSET = _td(hours=5, minutes=30)  # UTC+5:30

_VISITOR_LOG_FILE = Path(os.environ.get("VISITOR_LOG_PATH", "/tmp/visitor_logs.json" if CLOUD_MODE else str(APP_DIR / "visitor_logs.json")))

def _load_visitor_logs():
    try:
        if _VISITOR_LOG_FILE.exists():
            with open(_VISITOR_LOG_FILE, "r", encoding="utf-8") as f:
                return _json.load(f)
    except Exception:
        pass
    return []

def _save_visitor_logs(logs):
    try:
        with open(_VISITOR_LOG_FILE, "w", encoding="utf-8") as f:
            _json.dump(logs[:500], f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ Visitor log save failed: {e}")

@app.post("/api/track")
async def track_visit(request: Request):
    """Log public page visits — called from public pages."""
    body = await request.json()
    ip = request.headers.get("x-forwarded-for", request.client.host if request.client else "unknown")
    if "," in ip:
        ip = ip.split(",")[0].strip()
    ua = request.headers.get("user-agent", "")
    ua_lower = ua.lower()
    # Detect device brand
    if "iphone" in ua_lower or "ipad" in ua_lower:
        device = "iPhone"
    elif "samsung" in ua_lower:
        device = "Samsung"
    elif "xiaomi" in ua_lower or "redmi" in ua_lower or "poco" in ua_lower:
        device = "Xiaomi"
    elif "oppo" in ua_lower:
        device = "Oppo"
    elif "vivo" in ua_lower:
        device = "Vivo"
    elif "realme" in ua_lower:
        device = "Realme"
    elif "oneplus" in ua_lower:
        device = "OnePlus"
    elif "android" in ua_lower or "mobile" in ua_lower:
        device = "Mobile"
    else:
        device = "Desktop"
    ist_now = _dt.utcnow() + _IST_OFFSET
    entry = {
        "time": ist_now.strftime("%d %b, %I:%M %p"),
        "timestamp": ist_now.isoformat(),
        "page": body.get("page", "--"),
        "executive": body.get("executive", "--"),
        "device": device,
        "ip": ip[:20],
    }
    logs = _load_visitor_logs()
    logs.insert(0, entry)
    _save_visitor_logs(logs)
    return {"ok": True}

@app.get("/api/visitor-logs")
async def get_visitor_logs():
    """Admin endpoint — returns recent visitor logs."""
    logs = _load_visitor_logs()
    return {"logs": logs[:100]}

# ── PUBLIC LINKS ACCESS CONTROL ──
_public_access = {"enabled": True, "password_required": False, "password": "rcc123", "show_projection": True}
_search_access = {"enabled": False, "password": "rcc@admin"}

@app.get("/api/public-access")
async def get_public_access():
    return {
        "enabled": _public_access["enabled"],
        "password_required": _public_access["password_required"],
        "password": _public_access["password"],
        "show_projection": _public_access["show_projection"],
        "search_enabled": _search_access["enabled"],
        "search_password": _search_access["password"]
    }

@app.post("/api/public-access")
async def set_public_access(request: Request):
    body = await request.json()
    if "enabled" in body:
        _public_access["enabled"] = body["enabled"]
    if "password_required" in body:
        _public_access["password_required"] = body["password_required"]
    if "password" in body:
        _public_access["password"] = body["password"]
    if "show_projection" in body:
        _public_access["show_projection"] = body["show_projection"]
    if "search_enabled" in body:
        _search_access["enabled"] = body["search_enabled"]
    if "search_password" in body:
        _search_access["password"] = body["search_password"]
    return {
        "enabled": _public_access["enabled"],
        "password_required": _public_access["password_required"],
        "password": _public_access["password"],
        "show_projection": _public_access["show_projection"],
        "search_enabled": _search_access["enabled"],
        "search_password": _search_access["password"]
    }

@app.post("/api/public-verify")
async def verify_public_password(request: Request):
    body = await request.json()
    entered = body.get("password", "")
    verify_type = body.get("type", "general")
    if verify_type == "search":
        if entered == _search_access["password"]:
            return {"verified": True}
        return {"verified": False}
    if not _public_access["password_required"]:
        return {"verified": True}
    if entered == _public_access["password"]:
        return {"verified": True}
    return {"verified": False}

# ── PUBLIC PAGES (no login required) — must be before StaticFiles mount ──
from fastapi.responses import HTMLResponse

@app.get("/public/trails")
async def public_trails_page():
    if not _public_access["enabled"]:
        return HTMLResponse(content="""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>Access Disabled</title><style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Inter',sans-serif;background:#f0f4f8;display:flex;justify-content:center;align-items:center;min-height:100vh;padding:20px}.card{background:white;border-radius:16px;padding:40px;text-align:center;box-shadow:0 4px 20px rgba(0,0,0,0.08);max-width:360px}.icon{font-size:48px;margin-bottom:16px}.title{font-size:18px;font-weight:700;color:#1e293b;margin-bottom:8px}.msg{font-size:14px;color:#64748b}</style></head><body><div class="card"><div class="icon">🔒</div><div class="title">Access Disabled</div><div class="msg">This link has been disabled by the admin. Contact your administrator for access.</div></div></body></html>""", status_code=403)
    return FileResponse(str(APP_DIR / "mobile" / "public_trails.html"))

@app.get("/public/flowlist")
async def public_flowlist_page():
    if not _public_access["enabled"]:
        return HTMLResponse(content="""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>Access Disabled</title><style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Inter',sans-serif;background:#f0f4f8;display:flex;justify-content:center;align-items:center;min-height:100vh;padding:20px}.card{background:white;border-radius:16px;padding:40px;text-align:center;box-shadow:0 4px 20px rgba(0,0,0,0.08);max-width:360px}.icon{font-size:48px;margin-bottom:16px}.title{font-size:18px;font-weight:700;color:#1e293b;margin-bottom:8px}.msg{font-size:14px;color:#64748b}</style></head><body><div class="card"><div class="icon">🔒</div><div class="title">Access Disabled</div><div class="msg">This link has been disabled by the admin. Contact your administrator for access.</div></div></body></html>""", status_code=403)
    return FileResponse(str(APP_DIR / "mobile" / "public_flowlist.html"))

@app.get("/public/search")
async def public_search_page():
    if not _search_access["enabled"]:
        return HTMLResponse(content="""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>Access Disabled</title><style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Inter',sans-serif;background:#f0f4f8;display:flex;justify-content:center;align-items:center;min-height:100vh;padding:20px}.card{background:white;border-radius:16px;padding:40px;text-align:center;box-shadow:0 4px 20px rgba(0,0,0,0.08);max-width:360px}.icon{font-size:48px;margin-bottom:16px}.title{font-size:18px;font-weight:700;color:#1e293b;margin-bottom:8px}.msg{font-size:14px;color:#64748b}</style></head><body><div class="card"><div class="icon">🔒</div><div class="title">Access Disabled</div><div class="msg">Action Center is currently disabled. Contact your administrator.</div></div></body></html>""", status_code=403)
    return FileResponse(str(APP_DIR / "mobile" / "public_search.html"))

# ── FORCE SYNC API (admin can trigger manual refresh) ──
@app.post("/api/force-sync")
async def force_sync():
    """Force re-download from OneDrive and reload data."""
    if CLOUD_MODE and ONEDRIVE_SHARE_URL:
        success = _sync_from_onedrive()
        if success:
            _cache["df"] = None  # Force reload
            _cache["mtime"] = 0
            load_data()
            return {"status": "ok", "message": "Data synced from OneDrive"}
        return {"status": "error", "message": "OneDrive download failed"}
    else:
        _cache["df"] = None
        _cache["mtime"] = 0
        load_data()
        return {"status": "ok", "message": "Data reloaded from file"}


# ── TILL TIME REPORT API ──
@app.get("/api/report/tilltime")
async def tilltime_report(date: str = ""):
    """Generate till-time payment report data for today."""
    df = load_data()
    if df is None:
        return {"error": "Data not found"}
    
    if date:
        today = date.strip()
    else:
        today = _dt.now().strftime("%d.%m.%y") if not CLOUD_MODE else (_dt.utcnow() + _IST_OFFSET).strftime("%d.%m.%y")
    
    # Today's paid cases
    today_paid = df[(df["Payment Date"].astype(str).str.strip() == today) & 
                    (df["RECEIPT CUT"].astype(str).str.upper() == "PAID")]
    
    # Overall totals
    total_pos = float(df["POS"].sum())
    today_pos = float(today_paid["POS"].sum())
    
    # Receipt Cut Movement (today's paid / total cases %)
    rc_movement = round(len(today_paid) / len(df) * 100, 2) if len(df) else 0
    
    # BKT-1 Resolution Movement
    bkt1_total_pos = float(df[df["BUCKET"] == 1]["POS"].sum())
    bkt1_today_pos = float(today_paid[today_paid["BUCKET"] == 1]["POS"].sum())
    bkt1_movement = round(bkt1_today_pos / bkt1_total_pos * 100, 2) if bkt1_total_pos else 0
    
    # BKT-2 Resolution Movement
    bkt2_total_pos = float(df[df["BUCKET"] == 2]["POS"].sum())
    bkt2_today_pos = float(today_paid[today_paid["BUCKET"] == 2]["POS"].sum())
    bkt2_movement = round(bkt2_today_pos / bkt2_total_pos * 100, 2) if bkt2_total_pos else 0
    
    # Team wise count (BKT wise)
    team_count = {}
    team_pos = {}
    buckets_used = sorted(today_paid["BUCKET"].unique().tolist())
    
    for team, grp in today_paid.groupby("TEAM"):
        team_count[team] = {}
        team_pos[team] = {}
        for bkt, bgrp in grp.groupby("BUCKET"):
            team_count[team][int(bkt)] = len(bgrp)
            team_pos[team][int(bkt)] = round(float(bgrp["POS"].sum()), 2)
    
    # Grand totals per bucket
    grand_count = {}
    grand_pos = {}
    for bkt in buckets_used:
        bkt_data = today_paid[today_paid["BUCKET"] == bkt]
        grand_count[int(bkt)] = len(bkt_data)
        grand_pos[int(bkt)] = round(float(bkt_data["POS"].sum()), 2)
    
    return {
        "date": today,
        "total_paid_today": len(today_paid),
        "rc_movement": rc_movement,
        "bkt1_movement": bkt1_movement,
        "bkt2_movement": bkt2_movement,
        "buckets": [int(b) for b in buckets_used],
        "team_count": team_count,
        "team_pos": team_pos,
        "grand_count": grand_count,
        "grand_pos": grand_pos,
    }


# ── DAILY WINNERS API ──
@app.get("/api/report/winners")
async def daily_winners(date: str = ""):
    """Generate daily winners text for WhatsApp."""
    df = load_data()
    if df is None:
        return {"text": "❌ Data not found"}
    
    if date:
        today = date.strip()
    else:
        today = _dt.now().strftime("%d.%m.%y") if not CLOUD_MODE else (_dt.utcnow() + _IST_OFFSET).strftime("%d.%m.%y")
    
    # Today's paid cases
    today_paid = df[(df["Payment Date"].astype(str).str.strip() == today) & 
                    (df["RECEIPT CUT"].astype(str).str.upper() == "PAID")]
    
    if today_paid.empty:
        return {"text": "⭐ *TODAY'S WINNERS* 🏅\n\nNo payments yet today."}
    
    lines = ["⭐ *TODAY'S WINNERS* 🏅\n"]
    
    # BKT 1-2 Receipt Cut (2+ receipts = ₹150 per receipt)
    bkt12 = today_paid[today_paid["BUCKET"].isin([1, 2])]
    if not bkt12.empty:
        rc_by_team = bkt12.groupby("TEAM").size()
        rc_winners = rc_by_team[rc_by_team >= 2]
        if not rc_winners.empty:
            lines.append("*DAILY (BKT 1-2) RECEIPTS*")
            for team, count in rc_winners.sort_values(ascending=False).items():
                incentive = count * 150
                lines.append(f"{team} - {count} 💵 {incentive}")
            lines.append("")
    
    # BKT 1 — ₹50K+ POS (₹100 for 50K-1.99L, ₹200 for 2L-2.99L, ₹300 for 3L+ etc.)
    bkt1_paid = today_paid[today_paid["BUCKET"] == 1]
    if not bkt1_paid.empty:
        pos_by_team_b1 = bkt1_paid.groupby("TEAM")["POS"].sum()
        pos_winners_b1 = pos_by_team_b1[pos_by_team_b1 >= 50000]
        if not pos_winners_b1.empty:
            lines.append("*BUCKET 1 | ₹50K+ POS* 💰")
            for team, pos in pos_winners_b1.sort_values(ascending=False).items():
                if pos >= 200000:
                    incentive = int(pos // 100000) * 100
                    label = f">{int(pos/100000)}L"
                else:
                    incentive = 100
                    label = ">50k"
                lines.append(f"{team} {label} 💸{incentive}")
            lines.append("")
    
    # BKT 2 — ₹50K+ POS (same logic)
    bkt2_paid = today_paid[today_paid["BUCKET"] == 2]
    if not bkt2_paid.empty:
        pos_by_team_b2 = bkt2_paid.groupby("TEAM")["POS"].sum()
        pos_winners_b2 = pos_by_team_b2[pos_by_team_b2 >= 50000]
        if not pos_winners_b2.empty:
            lines.append("*BUCKET 2 | ₹50K+ POS* 💰")
            for team, pos in pos_winners_b2.sort_values(ascending=False).items():
                if pos >= 200000:
                    incentive = int(pos // 100000) * 100
                    label = f">{int(pos/100000)}L"
                else:
                    incentive = 100
                    label = ">50k"
                lines.append(f"{team} {label} 💸{incentive}")
            lines.append("")
    
    # BKT 3-6 Receipt Cut (₹100 per receipt, no minimum)
    bkt36 = today_paid[today_paid["BUCKET"].isin([3, 4, 5, 6])]
    if not bkt36.empty:
        rc_by_team_36 = bkt36.groupby("TEAM").size()
        if not rc_by_team_36.empty:
            lines.append("*DAILY (BKT 3-6) RECEIPTS*")
            for team, count in rc_by_team_36.sort_values(ascending=False).items():
                incentive = count * 100
                lines.append(f"{team} - {count} 💵 {incentive}")
            lines.append("")
    
    # SUNDAY SPECIAL — ₹200 per receipt (any bucket)
    try:
        day_parts = today.split(".")
        check_date = _dt(2000 + int(day_parts[2]), int(day_parts[1]), int(day_parts[0]))
        is_sunday = check_date.weekday() == 6  # 6 = Sunday
    except Exception:
        is_sunday = False
    
    if is_sunday and not today_paid.empty:
        sunday_by_team = today_paid.groupby("TEAM").size()
        if not sunday_by_team.empty:
            lines.append("*🔴 SUNDAY SPECIAL (₹200/receipt)*")
            for team, count in sunday_by_team.sort_values(ascending=False).items():
                incentive = count * 200
                lines.append(f"{team} - {count} 💵 {incentive}")
            lines.append("")
    
    return {"text": "\n".join(lines)}


# ── RESOLUTION TABLE API ──
@app.get("/api/report/resolution")
async def resolution_table(bucket: int = 1):
    """Generate resolution table data for specified bucket — POS STATUS wise."""
    df = load_data()
    if df is None:
        return {"error": "Data not found"}
    
    bkt_df = df[df["BUCKET"] == bucket]
    if bkt_df.empty:
        return {"bucket": bucket, "teams": [], "grand": {}, "movement": 0}
    
    # Today's movement (today's paid POS / total POS for this bucket)
    today = _dt.now().strftime("%d.%m.%y") if not CLOUD_MODE else (_dt.utcnow() + _IST_OFFSET).strftime("%d.%m.%y")
    today_paid_bkt = bkt_df[(bkt_df["Payment Date"].astype(str).str.strip() == today) & (bkt_df["RECEIPT CUT"].astype(str).str.upper() == "PAID")]
    total_pos = float(bkt_df["POS"].sum())
    today_pos = float(today_paid_bkt["POS"].sum())
    movement = round(today_pos / total_pos * 100, 2) if total_pos else 0
    
    # Team wise POS by POS STATUS
    teams = []
    for team, grp in bkt_df.groupby("TEAM"):
        flow_pos = float(grp[grp["POS STATUS"] == "FLOW"]["POS"].sum())
        stable_pos = float(grp[grp["POS STATUS"] == "STABLE"]["POS"].sum())
        rb_pos = float(grp[grp["POS STATUS"] == "RB"]["POS"].sum())
        grand_total = flow_pos + stable_pos + rb_pos
        stable_pct = round(stable_pos / grand_total * 100, 2) if grand_total else 0
        rb_pct = round(rb_pos / grand_total * 100, 2) if grand_total else 0
        resl = round(stable_pct + rb_pct, 2)
        flow_pct = round(flow_pos / total_pos * 100, 2) if total_pos else 0
        flow_cases = int((grp["POS STATUS"] == "FLOW").sum())
        # Current resolution from POS STATUS
        current_res = round((stable_pos + rb_pos) / grand_total * 100, 2) if grand_total else 0
        # Agency % — average of AGENCY CASE% for this team (stored as decimal)
        agency_pct = 0
        if "AGENCY CASE%" in grp.columns:
            agency_vals = grp["AGENCY CASE%"].dropna()
            if len(agency_vals) > 0:
                agency_pct = round(float(agency_vals.mean()) * 100, 2)
        # Projection resolution from PROJECTION column
        proj_resl = 0
        if "PROJECTION" in grp.columns:
            proj_grp = grp.groupby("PROJECTION")["POS"].sum()
            p_stable = float(proj_grp.get("STABLE", 0))
            p_rb = float(proj_grp.get("RB", 0))
            p_total = float(proj_grp.sum())
            proj_resl = round((p_stable + p_rb) / p_total * 100, 2) if p_total else 0
        teams.append({
            "team": str(team),
            "flow": round(flow_pos, 1),
            "stable": round(stable_pos, 1),
            "rb": round(rb_pos, 1),
            "grand_total": round(grand_total, 1),
            "stable_pct": stable_pct,
            "rb_pct": rb_pct,
            "resl": resl,
            "flow_pct": flow_pct,
            "flow_cases": flow_cases,
            "agency_pct": agency_pct,
            "current_res": current_res,
            "proj_resl": proj_resl,
        })
    
    teams.sort(key=lambda x: x["resl"], reverse=True)
    
    # Grand total
    total_flow = sum(t["flow"] for t in teams)
    total_stable = sum(t["stable"] for t in teams)
    total_rb = sum(t["rb"] for t in teams)
    total_all = total_flow + total_stable + total_rb
    # Grand agency average
    grand_agency = round(sum(t["agency_pct"] for t in teams) / len(teams), 2) if teams else 0
    grand_flow_cases = sum(t["flow_cases"] for t in teams)
    grand_flow_pct = round(total_flow / total_all * 100, 2) if total_all else 0
    grand = {
        "flow": round(total_flow, 1),
        "stable": round(total_stable, 1),
        "rb": round(total_rb, 1),
        "grand_total": round(total_all, 1),
        "stable_pct": round(total_stable / total_all * 100, 2) if total_all else 0,
        "rb_pct": round(total_rb / total_all * 100, 2) if total_all else 0,
        "resl": round((total_stable + total_rb) / total_all * 100, 2) if total_all else 0,
        "flow_cases": grand_flow_cases,
        "flow_pct": grand_flow_pct,
        "agency_pct": grand_agency,
    }
    
    return {"bucket": bucket, "movement": movement, "teams": teams, "grand": grand}


# ── RECEIPT CUT REPORT API ──
@app.get("/api/report/receiptcut")
async def receiptcut_report():
    """Generate Receipt Cut report — PAID/UNPAID/Target/Shortfall per executive."""
    df = load_data()
    if df is None:
        return {"error": "Data not found"}
    
    # Determine target based on current date
    ist_now = _dt.now() if not CLOUD_MODE else _dt.utcnow() + _IST_OFFSET
    day = ist_now.day
    if day <= 10:
        target_pct = 25
    elif day <= 20:
        target_pct = 45
    else:
        target_pct = 65
    
    # Days remaining in current range
    if day <= 10:
        days_remaining = 10 - day + 1
    elif day <= 20:
        days_remaining = 20 - day + 1
    else:
        import calendar
        last_day = calendar.monthrange(ist_now.year, ist_now.month)[1]
        days_remaining = last_day - day + 1
    
    # Today's data
    today = ist_now.strftime("%d.%m.%y")
    today_paid = df[(df["Payment Date"].astype(str).str.strip() == today) & (df["RECEIPT CUT"].astype(str).str.upper() == "PAID")]
    total_pos = float(df["POS"].sum())
    today_pos = float(today_paid["POS"].sum())
    movement = round(today_pos / total_pos * 100, 2) if total_pos else 0
    
    # Team wise stats
    teams = []
    for team, grp in df.groupby("TEAM"):
        total = len(grp)
        paid = int((grp["RECEIPT CUT"].astype(str).str.upper() == "PAID").sum())
        unpaid = total - paid
        target = int(round(total * target_pct / 100))
        shortfall = max(0, target - paid)
        drr = round(shortfall / days_remaining, 0) if days_remaining > 0 else 0
        pct_achi = round(paid / total * 100, 2) if total else 0
        # Today's payment count for this team
        today_payment = int(today_paid[today_paid["TEAM"] == team].shape[0]) if not today_paid.empty else 0
        # Today's trails
        today_trails = 0  # Would need CSV data — skip for now
        # Pending trails
        pending_trails = int((grp["TRAILS PENDING"] == 0).sum())
        
        teams.append({
            "team": str(team),
            "paid": paid,
            "unpaid": unpaid,
            "total": total,
            "target": target,
            "shortfall": shortfall,
            "drr": int(drr),
            "pct_achi": pct_achi,
            "payment": today_payment,
            "today_trails": today_trails,
            "pending_trails": pending_trails,
        })
    
    teams.sort(key=lambda x: x["pct_achi"], reverse=True)
    
    # Grand total
    g_paid = sum(t["paid"] for t in teams)
    g_unpaid = sum(t["unpaid"] for t in teams)
    g_total = sum(t["total"] for t in teams)
    g_target = int(round(g_total * target_pct / 100))
    g_shortfall = max(0, g_target - g_paid)
    g_drr = int(round(g_shortfall / days_remaining)) if days_remaining > 0 else 0
    g_pct = round(g_paid / g_total * 100, 2) if g_total else 0
    g_payment = sum(t["payment"] for t in teams)
    g_pending = sum(t["pending_trails"] for t in teams)
    
    grand = {
        "paid": g_paid, "unpaid": g_unpaid, "total": g_total,
        "target": g_target, "shortfall": g_shortfall, "drr": g_drr,
        "pct_achi": g_pct, "payment": g_payment, "pending_trails": g_pending,
    }
    
    return {
        "movement": movement,
        "target_pct": target_pct,
        "days_remaining": days_remaining,
        "day": day,
        "teams": teams,
        "grand": grand,
    }


# ── RECEIPT CUT REPORT API ──
@app.get("/api/report/receiptcut")
async def receipt_cut_report():
    """Generate receipt cut report with date-based targets."""
    df = load_data()
    if df is None:
        return {"error": "Data not found"}
    
    # Determine target based on current date
    today_dt = _dt.now() if not CLOUD_MODE else _dt.utcnow() + _IST_OFFSET
    day = today_dt.day
    if day <= 10:
        target_pct = 25
        remaining_days = 10 - day + 1
    elif day <= 20:
        target_pct = 45
        remaining_days = 20 - day + 1
    else:
        target_pct = 65
        last_day = 30  # approximate
        remaining_days = max(last_day - day + 1, 1)
    
    # Today's movement
    today = today_dt.strftime("%d.%m.%y")
    today_paid = df[(df["Payment Date"].astype(str).str.strip() == today) & (df["RECEIPT CUT"].astype(str).str.upper() == "PAID")]
    total_cases = len(df)
    movement = round(len(today_paid) / total_cases * 100, 2) if total_cases else 0
    
    # Team wise data
    teams = []
    for team, grp in df.groupby("TEAM"):
        total = len(grp)
        paid = int((grp["RECEIPT CUT"].astype(str).str.upper() == "PAID").sum())
        unpaid = total - paid
        target_cases = int(round(total * target_pct / 100))
        shortfall = target_cases - paid
        drr = int(round(shortfall / remaining_days)) if remaining_days > 0 and shortfall > 0 else 0
        pct_achi = round(paid / total * 100, 2) if total else 0
        # Today's payment count for this team
        today_payment = int(today_paid[today_paid["TEAM"] == team].shape[0]) if not today_paid.empty else 0
        # Today's trails
        today_trails = 0  # Would need trails data
        # Pending trails
        pending_trails = int((grp["TRAILS PENDING"] == 0).sum())
        
        teams.append({
            "team": str(team),
            "paid": paid,
            "unpaid": unpaid,
            "total": total,
            "target": target_cases,
            "shortfall": shortfall,
            "drr": int(drr),
            "pct_achi": pct_achi,
            "payment": today_payment,
            "pending_trails": pending_trails,
        })
    
    teams.sort(key=lambda x: x["pct_achi"], reverse=True)
    
    # Grand total
    grand = {
        "paid": sum(t["paid"] for t in teams),
        "unpaid": sum(t["unpaid"] for t in teams),
        "total": sum(t["total"] for t in teams),
        "target": round(sum(t["target"] for t in teams), 2),
        "shortfall": round(sum(t["shortfall"] for t in teams), 2),
        "drr": round(sum(t["drr"] for t in teams), 2),
        "pct_achi": round(sum(t["paid"] for t in teams) / sum(t["total"] for t in teams) * 100, 2) if sum(t["total"] for t in teams) else 0,
        "payment": sum(t["payment"] for t in teams),
        "pending_trails": sum(t["pending_trails"] for t in teams),
    }
    
    return {
        "movement": movement,
        "target_pct": target_pct,
        "remaining_days": remaining_days,
        "day": day,
        "teams": teams,
        "grand": grand,
    }


# ── TRAILS CSV UPLOAD API ──
from fastapi import UploadFile, File
import csv
import io

@app.post("/api/trails/upload")
async def upload_trails_csv(file: UploadFile = File(...)):
    """Upload Vymo trails CSV → match loan nos → update TRAILS PENDING → return report."""
    try:
        content = await file.read()
        text = content.decode("utf-8")
        
        # Parse CSV — extract loan nos from "Row Labels" column
        loan_nos = []
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if not row:
                continue
            val = row[0].strip()
            # Skip headers and "Grand Total"
            if val in ("Row Labels", "Contact Mode", "", "Grand Total") or not val.isdigit():
                continue
            loan_nos.append(val)
        
        if not loan_nos:
            return {"error": "No loan numbers found in CSV", "matched": 0}
        
        # Load current data
        df = load_data()
        if df is None:
            return {"error": "Data file not found"}
        
        # Match loan nos → get TEAM
        matched = df[df["LOAN NO"].isin(loan_nos)]
        team_count = matched.groupby("TEAM").size().sort_values(ascending=False).to_dict()
        total_matched = len(matched)
        total_csv = len(loan_nos)
        unmatched = total_csv - total_matched
        
        # Pending count (TRAILS PENDING = 0 per team)
        pending_df = df[df["TRAILS PENDING"] == 0]
        pending_count = pending_df.groupby("TEAM").size().to_dict()
        total_pending = len(pending_df)
        
        # Update TRAILS PENDING in Excel (only on local PC, not cloud/Render)
        if not CLOUD_MODE:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(DATA_FILE))
                ws = wb["MAIN"]
                
                # Find LOAN NO and TRAILS PENDING column indices
                headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
                loan_col = headers.index("LOAN NO") + 1 if "LOAN NO" in headers else None
                trail_col = headers.index("TRAILS PENDING") + 1 if "TRAILS PENDING" in headers else None
                
                if loan_col and trail_col:
                    updated = 0
                    for row_idx in range(2, ws.max_row + 1):
                        cell_val = str(ws.cell(row=row_idx, column=loan_col).value or "").strip()
                        if cell_val in loan_nos:
                            current = ws.cell(row=row_idx, column=trail_col).value or 0
                            try:
                                current = int(current)
                            except (ValueError, TypeError):
                                current = 0
                            ws.cell(row=row_idx, column=trail_col, value=current + 1)
                            updated += 1
                    
                    wb.save(str(DATA_FILE))
                    wb.close()
                    
                    # Clear cache so next load picks up changes
                    _cache["df"] = None
                    _cache["mtime"] = 0
                    
                    # Also copy to OneDrive
                    import shutil
                    onedrive_path = Path(r"C:\Users\BAJAJ1\OneDrive\RCC\RCC_DATA.xlsx")
                    if onedrive_path.parent.exists():
                        shutil.copy2(str(DATA_FILE), str(onedrive_path))
            except Exception as e:
                print(f"⚠️ Excel update failed: {e}")
        
        # Generate WhatsApp report text
        today_str = _dt.now().strftime("%d %b")
        lines = [f"📋 *TRAILS REPORT* — {today_str}\n"]
        lines.append(f"*TODAY'S TRAILS DONE ({total_matched})*")
        for team, count in sorted(team_count.items()):
            lines.append(f"{team} — {count}")
        lines.append(f"*TOTAL — {total_matched}*\n")
        
        lines.append(f"*PENDING TRAILS ({total_pending})*")
        for team in sorted(pending_count.keys()):
            lines.append(f"{team} — {pending_count[team]}")
        lines.append(f"*TOTAL — {total_pending}*")
        
        return {
            "text": "\n".join(lines),
            "matched": total_matched,
            "unmatched": unmatched,
            "team_count": team_count,
            "pending_count": pending_count,
            "total_pending": total_pending,
        }
    except Exception as e:
        return {"error": str(e), "matched": 0}


app.mount("/mobile", StaticFiles(directory=str(APP_DIR / "mobile"), html=True), name="mobile")


import time as _time
_last_onedrive_check = {"t": 0}

def load_data():
    """Load and clean Excel data with caching — only re-reads when file changes."""
    # In cloud mode, check OneDrive every 5 minutes for fresh data
    if CLOUD_MODE and ONEDRIVE_SHARE_URL:
        now = _time.time()
        if now - _last_onedrive_check["t"] > 120:  # 2 min
            _last_onedrive_check["t"] = now
            _sync_from_onedrive()

    if not DATA_FILE.exists():
        return None
    
    # Check file modification time
    current_mtime = os.path.getmtime(DATA_FILE)
    if _cache["df"] is not None and _cache["mtime"] == current_mtime:
        return _cache["df"]
    
    print(f"📂 Loading Excel (file changed)...")
    # Load only required columns to save memory (512MB limit on Render)
    needed_cols = ["LOAN NO", "CUSTOMER NAME", "Mode Of Payment", "Payment Date", 
                   "Paid Amount", "EMI", "TOTAL EMI DUE", "STAB AMOUNTWITH DPIC", 
                   "RB AMOUNTWITH DPIC", "BUCKET", "POS STATUS", "RECEIPT CUT", 
                   "TEAM", "POS", "DPIC CHARGES", "DRA CASE%", "AGENCY CASE%", 
                   "AREA", "MOBILE", "DPD", "TRAILS PENDING", "PROJECTION"]
    try:
        df = pd.read_excel(DATA_FILE, sheet_name=SHEET_NAME, engine="openpyxl", usecols=lambda c: str(c).strip() in needed_cols)
    except Exception:
        df = pd.read_excel(DATA_FILE, sheet_name=SHEET_NAME, engine="openpyxl")
    df.columns = [str(col).strip() for col in df.columns]
    
    num_cols = ["BUCKET", "POS", "EMI", "TOTAL EMI DUE", "DPIC CHARGES",
                "STAB AMOUNTWITH DPIC", "RB AMOUNTWITH DPIC", "Paid Amount",
                "TRAILS PENDING", "DRA CASE%", "AGENCY CASE%", "DPD"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(",", "").str.replace("₹", "").str.strip(),
                errors="coerce"
            ).fillna(0)
    
    if "LOAN NO" in df.columns:
        df["LOAN NO"] = df["LOAN NO"].astype(str).str.strip()
    if "CUSTOMER NAME" in df.columns:
        df["CUSTOMER NAME"] = df["CUSTOMER NAME"].astype(str).str.strip()
    if "TEAM" in df.columns:
        df["TEAM"] = df["TEAM"].astype(str).str.strip().str.upper()
    if "POS STATUS" in df.columns:
        df["POS STATUS"] = df["POS STATUS"].astype(str).str.strip().str.upper()
    if "BUCKET" in df.columns:
        df["BUCKET"] = pd.to_numeric(
            df["BUCKET"].astype(str).str.extract(r"(\d+)", expand=False),
            errors="coerce"
        ).fillna(0).astype(int)
    if "RECEIPT CUT" in df.columns:
        df["RECEIPT CUT"] = df["RECEIPT CUT"].astype(str).str.strip().str.upper()
    if "AREA" in df.columns:
        df["AREA"] = df["AREA"].astype(str).str.strip()
    
    # Store in cache
    _cache["df"] = df
    _cache["mtime"] = current_mtime
    print(f"✅ Cached! {len(df)} rows loaded.")
    
    # Free memory
    import gc
    gc.collect()
    
    return df


def resolution_stats(df):
    """Calculate resolution stats from dataframe."""
    pos = df["POS"]
    status = df["POS STATUS"]
    flow_c = int((status == "FLOW").sum())
    stable_c = int((status == "STABLE").sum())
    rb_c = int((status == "RB").sum())
    flow_p = float(pos[status == "FLOW"].sum())
    stable_p = float(pos[status == "STABLE"].sum())
    rb_p = float(pos[status == "RB"].sum())
    total_p = flow_p + stable_p + rb_p
    res = ((stable_p + rb_p) / total_p * 100) if total_p else 0
    return {
        "flow": flow_c, "stable": stable_c, "rb": rb_c,
        "flow_pos": flow_p, "stable_pos": stable_p, "rb_pos": rb_p,
        "total_pos": total_p, "resolution_pct": round(res, 2)
    }


def calc_payout_slab(resolution_pct, rb_pct, bucket):
    """Calculate payout slab."""
    if bucket not in (1, 2):
        return 8, True
    
    PAYOUT_CONFIG = {
        1: [(90, 999, 15, 25), (88, 89.99, 12, 20), (85, 87.99, 10, 15), (0, 84.99, 8, 15)],
        2: [(70, 999, 15, 25), (65, 69.99, 12, 20), (60, 64.99, 10, 15), (0, 59.99, 8, 15)],
    }
    
    slabs = PAYOUT_CONFIG.get(bucket, PAYOUT_CONFIG[2])
    payout = 8
    rb_target = 15
    
    for min_res, max_res, rate, rb_req in slabs:
        if min_res <= resolution_pct <= max_res:
            payout = rate
            rb_target = rb_req
            break
    
    rb_achieved = rb_pct >= rb_target
    if not rb_achieved:
        payout = max(8, payout - 2)
    
    return payout, rb_achieved


# ── API ENDPOINTS ──

@app.post("/api/login")
async def login(request: Request):
    body = await request.json()
    username = body.get("username", "").strip().upper()
    password = body.get("password", "")
    
    rec = CREDENTIALS.get(username)
    if not rec or rec["password"] != password:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    return {"username": username, "role": rec["role"]}


@app.get("/api/dashboard")
async def dashboard(user: str = "", role: str = "executive"):
    df = load_data()
    if df is None:
        raise HTTPException(status_code=500, detail="Data file not found")
    
    username = user.strip().upper()
    
    # Filter by user if executive
    if role != "admin" and username:
        exec_df = df[df["TEAM"] == username].copy()
    else:
        exec_df = df.copy()
    
    total_cases = len(exec_df)
    if total_cases == 0:
        return {"total_cases": 0, "error": "No data found"}
    
    # Overall stats
    paid_count = int((exec_df["RECEIPT CUT"] == "PAID").sum())
    unpaid_count = total_cases - paid_count
    rc_pct = round(paid_count / total_cases * 100, 2) if total_cases else 0
    total_collection = float(exec_df["Paid Amount"].sum())
    
    # Bucket stats
    def bkt_stats(data, bkt):
        g = data[data["BUCKET"] == bkt]
        total = len(g)
        if total == 0:
            return {"total": 0, "flow": 0, "stable": 0, "rb": 0,
                    "res_pct": 0, "rb_pct": 0, "collection": 0, "slab": 8, "payout": 0}
        flow = int((g["POS STATUS"] == "FLOW").sum())
        stable = int((g["POS STATUS"] == "STABLE").sum())
        rb = int((g["POS STATUS"] == "RB").sum())
        
        # Calculate resolution & RB% using POS amounts (not case counts)
        total_pos = float(g["POS"].sum())
        stable_pos = float(g[g["POS STATUS"] == "STABLE"]["POS"].sum())
        rb_pos = float(g[g["POS STATUS"] == "RB"]["POS"].sum())
        
        if total_pos > 0:
            res_pct = round((stable_pos + rb_pos) / total_pos * 100, 2)
            rb_pct = round(rb_pos / total_pos * 100, 2)
        else:
            res_pct = 0
            rb_pct = 0
        
        collection = float(g["Paid Amount"].sum())
        slab, rb_ok = calc_payout_slab(res_pct, rb_pct, bkt)
        payout = collection * slab / 100
        return {"total": total, "flow": flow, "stable": stable, "rb": rb,
                "res_pct": res_pct, "rb_pct": rb_pct, "collection": collection,
                "slab": slab, "payout": round(payout, 2), "rb_ok": rb_ok}
    
    b1 = bkt_stats(exec_df, 1)
    b2 = bkt_stats(exec_df, 2)
    
    # BKT 3-6
    other_collection = float(exec_df[~exec_df["BUCKET"].isin([1, 2])]["Paid Amount"].sum())
    other_payout = other_collection * 8 / 100
    
    total_payout = b1["payout"] + b2["payout"] + other_payout
    
    # Slab projections
    slab_10 = total_collection * 0.10
    slab_12 = total_collection * 0.12
    slab_15 = total_collection * 0.15
    
    return {
        "total_cases": total_cases,
        "paid": paid_count,
        "unpaid": unpaid_count,
        "receipt_pct": rc_pct,
        "total_collection": round(total_collection, 2),
        "total_payout": round(total_payout, 2),
        "bkt1": b1,
        "bkt2": b2,
        "other_collection": round(other_collection, 2),
        "other_payout": round(other_payout, 2),
        "slab_10": round(slab_10, 2),
        "slab_12": round(slab_12, 2),
        "slab_15": round(slab_15, 2),
    }


@app.get("/api/trails")
async def trails(user: str = "", role: str = "executive", auth: str = ""):
    # Block if public access is disabled and no auth token
    if not _public_access["enabled"] and auth != "rcc-admin-token":
        raise HTTPException(status_code=403, detail="Access disabled")
    df = load_data()
    if df is None:
        raise HTTPException(status_code=500, detail="Data file not found")
    
    username = user.strip().upper()
    
    # Admin sees all pending trails, executive sees only their own
    if role == "admin":
        pending = df[df["TRAILS PENDING"] == 0]
    else:
        pending = df[(df["TEAM"] == username) & (df["TRAILS PENDING"] == 0)]
    
    result = []
    for _, row in pending.iterrows():
        result.append({
            "loan_no": str(row["LOAN NO"]),
            "customer_name": str(row["CUSTOMER NAME"]),
            "team": str(row["TEAM"]),
            "area": str(row.get("AREA", "—")),
        })
    
    return {"total": len(result), "trails": result}


@app.get("/api/flowlist")
async def flowlist(user: str = "", bucket: int = 1, role: str = "executive", auth: str = ""):
    # Block if public access is disabled and no auth token
    if not _public_access["enabled"] and auth != "rcc-admin-token":
        raise HTTPException(status_code=403, detail="Access disabled")
    try:
        df = load_data()
        if df is None:
            raise HTTPException(status_code=500, detail="Data file not found")
        
        username = user.strip().upper()
        
        # Admin sees all flow cases, executive sees only their own
        if role == "admin":
            flow_df = df[(df["POS STATUS"] == "FLOW") & (df["BUCKET"] == bucket)]
        else:
            flow_df = df[(df["TEAM"] == username) & (df["POS STATUS"] == "FLOW") & (df["BUCKET"] == bucket)]
        
        flow_df = flow_df.sort_values("POS", ascending=False)
        
        result = []
        for _, row in flow_df.iterrows():
            try:
                result.append({
                    "customer_name": str(row["CUSTOMER NAME"]) if "CUSTOMER NAME" in row.index else "",
                    "team": str(row["TEAM"]) if "TEAM" in row.index else "",
                    "pos": float(row["POS"]) if pd.notna(row.get("POS")) else 0,
                    "dra_pct": round(float(row["DRA CASE%"]) * 100, 1) if pd.notna(row.get("DRA CASE%")) else 0,
                    "mobile": str(row.get("MOBILE", "")),
                    "area": str(row.get("AREA", "")),
                    "projection": str(row["PROJECTION"]) if "PROJECTION" in row.index else "",
                })
            except Exception:
                continue
        
        try:
            proj = _calc_projection(df, bucket, username)
        except Exception:
            proj = {"resolution": 0, "current_res": 0, "current_rb": 0}
        
        return {"total": len(result), "bucket": bucket, "cases": result, "projection": proj}
    except HTTPException:
        raise
    except Exception as e:
        return {"total": 0, "bucket": bucket, "cases": [], "projection": {"resolution": 0, "current_res": 0, "current_rb": 0}, "error": str(e)}


def _calc_projection(df, bucket, username):
    """Calculate inline projection + current res for flowlist banner."""
    try:
        bkt_df = df[df["BUCKET"] == bucket]
        if username:
            user_bkt = bkt_df[bkt_df["TEAM"] == username]
        else:
            user_bkt = bkt_df
        if user_bkt.empty:
            return {"resolution": 0, "current_res": 0, "current_rb": 0}
        # Projection from PROJECTION column
        if "PROJECTION" in user_bkt.columns:
            proj_grp = user_bkt.groupby("PROJECTION")["POS"].sum()
            p_stable = float(proj_grp.get("STABLE", 0))
            p_rb = float(proj_grp.get("RB", 0))
            p_total = float(proj_grp.sum())
            resolution = round((p_stable + p_rb) / p_total * 100, 2) if p_total else 0
        else:
            resolution = 0
        # Current from POS STATUS
        cur_grp = user_bkt.groupby("POS STATUS")["POS"].sum()
        c_stable = float(cur_grp.get("STABLE", 0))
        c_rb = float(cur_grp.get("RB", 0))
        c_total = float(cur_grp.sum())
        current_res = round((c_stable + c_rb) / c_total * 100, 2) if c_total else 0
        current_rb = round(c_rb / c_total * 100, 2) if c_total else 0
        return {"resolution": resolution, "current_res": current_res, "current_rb": current_rb}
    except Exception:
        return {"resolution": 0, "current_res": 0, "current_rb": 0}


@app.get("/api/search")
async def search_loan(q: str = "", user: str = "", role: str = "executive", bucket: int = 0, status: str = ""):
    df = load_data()
    if df is None:
        raise HTTPException(status_code=500, detail="Data file not found")
    
    username = user.strip().upper()
    
    if role != "admin" and username:
        search_df = df[df["TEAM"] == username]
    else:
        search_df = df
    
    # Bucket filter
    if bucket > 0:
        search_df = search_df[search_df["BUCKET"] == bucket]
    
    # POS Status filter (FLOW / STABLE / RB)
    if status and status.upper() in ("FLOW", "STABLE", "RB"):
        search_df = search_df[search_df["POS STATUS"] == status.upper()]
    
    # If no query, return all cases sorted by STAB amount
    if not q.strip():
        results = search_df.sort_values("STAB AMOUNTWITH DPIC", ascending=False).head(50)
    else:
        query = q.strip().upper()
        mask = (
            search_df["LOAN NO"].str.startswith(query, na=False) |
            search_df["CUSTOMER NAME"].str.upper().str.contains(query, na=False)
        )
        results = search_df[mask].sort_values("STAB AMOUNTWITH DPIC", ascending=False).head(50)
    
    items = []
    for _, row in results.iterrows():
        items.append({
            "loan_no": str(row["LOAN NO"]),
            "customer_name": str(row["CUSTOMER NAME"]),
            "team": str(row["TEAM"]),
            "bucket": int(row["BUCKET"]),
            "pos": float(row["POS"]),
            "pos_status": str(row["POS STATUS"]),
            "emi": float(row["EMI"]),
            "emi_due": float(row["TOTAL EMI DUE"]),
            "stab_amount": float(row["STAB AMOUNTWITH DPIC"]),
            "rb_amount": float(row["RB AMOUNTWITH DPIC"]),
            "dpic": float(row["DPIC CHARGES"]),
            "receipt_cut": str(row["RECEIPT CUT"]),
            "area": str(row.get("AREA", "")),
            "mobile": str(row.get("MOBILE", "")),
            "dra_pct": round(float(row["DRA CASE%"]) * 100, 1),
            "paid_amount": float(row["Paid Amount"]),
            "dpd": int(row.get("DPD", 0)),
            "trails_pending": int(row.get("TRAILS PENDING", 0)),
        })
    
    return {"results": items}


@app.get("/api/executives")
async def executives():
    df = load_data()
    if df is None:
        raise HTTPException(status_code=500, detail="Data file not found")
    
    teams = sorted(df["TEAM"].dropna().unique().tolist())
    return {"executives": teams}


@app.get("/api/projection")
async def projection(bucket: int = 1, user: str = ""):
    """BKT-wise projection pivot — TEAM × PROJECTION(FLOW/STABLE/RB) with Resolution%."""
    df = load_data()
    if df is None:
        raise HTTPException(status_code=500, detail="Data file not found")
    
    bkt_df = df[df["BUCKET"] == bucket]
    
    # If specific user requested, filter for that user only
    username = user.strip().upper()
    if username:
        user_df = bkt_df[bkt_df["TEAM"] == username]
        if user_df.empty:
            return {"bucket": bucket, "user": username, "resolution": 0, "stable_pct": 0, "rb_pct": 0, "current_res": 0, "teams": [], "grand_total": {}}
        # Projection (from PROJECTION column)
        proj_grp = user_df.groupby("PROJECTION")["POS"].sum()
        flow_val = float(proj_grp.get("FLOW", 0))
        stable_val = float(proj_grp.get("STABLE", 0))
        rb_val = float(proj_grp.get("RB", 0))
        total_val = flow_val + stable_val + rb_val
        s_pct = round(stable_val / total_val * 100, 2) if total_val else 0
        r_pct = round(rb_val / total_val * 100, 2) if total_val else 0
        res = round(s_pct + r_pct, 2)
        # Current Resolution (from POS STATUS column — actual real-time)
        cur_grp = user_df.groupby("POS STATUS")["POS"].sum()
        cur_stable = float(cur_grp.get("STABLE", 0))
        cur_rb = float(cur_grp.get("RB", 0))
        cur_total = float(cur_grp.sum())
        cur_res = round((cur_stable + cur_rb) / cur_total * 100, 2) if cur_total else 0
        return {"bucket": bucket, "user": username, "resolution": res, "stable_pct": s_pct, "rb_pct": r_pct,
                "current_res": cur_res,
                "flow": round(flow_val, 2), "stable": round(stable_val, 2), "rb": round(rb_val, 2), "grand_total": round(total_val, 2)}
    
    if bkt_df.empty:
        return {"bucket": bucket, "teams": [], "grand_total": {}}
    
    # Group by TEAM and PROJECTION column (FLOW/STABLE/RB)
    pivot = bkt_df.groupby(["TEAM", "PROJECTION"])["POS"].sum().unstack(fill_value=0)
    
    # Ensure all columns exist
    for col in ["FLOW", "STABLE", "RB"]:
        if col not in pivot.columns:
            pivot[col] = 0.0
    
    pivot["grand_total"] = pivot["FLOW"] + pivot["STABLE"] + pivot["RB"]
    pivot["stable_pct"] = (pivot["STABLE"] / pivot["grand_total"] * 100).round(2).fillna(0)
    pivot["rb_pct"] = (pivot["RB"] / pivot["grand_total"] * 100).round(2).fillna(0)
    pivot["resolution"] = (pivot["stable_pct"] + pivot["rb_pct"]).round(2)
    
    # Sort by resolution descending
    pivot = pivot.sort_values("resolution", ascending=False)
    
    teams = []
    for team_name, row in pivot.iterrows():
        teams.append({
            "team": str(team_name),
            "flow": round(float(row["FLOW"]), 2),
            "stable": round(float(row["STABLE"]), 2),
            "rb": round(float(row["RB"]), 2),
            "grand_total": round(float(row["grand_total"]), 2),
            "stable_pct": float(row["stable_pct"]),
            "rb_pct": float(row["rb_pct"]),
            "resolution": float(row["resolution"]),
        })
    
    # Grand total row
    total_flow = pivot["FLOW"].sum()
    total_stable = pivot["STABLE"].sum()
    total_rb = pivot["RB"].sum()
    total_all = total_flow + total_stable + total_rb
    grand = {
        "flow": round(total_flow, 2),
        "stable": round(total_stable, 2),
        "rb": round(total_rb, 2),
        "grand_total": round(total_all, 2),
        "stable_pct": round(total_stable / total_all * 100, 2) if total_all else 0,
        "rb_pct": round(total_rb / total_all * 100, 2) if total_all else 0,
        "resolution": round((total_stable + total_rb) / total_all * 100, 2) if total_all else 0,
    }
    
    # Current Resolution from POS STATUS (actual real-time)
    cur_grp = bkt_df.groupby("POS STATUS")["POS"].sum()
    cur_stable = float(cur_grp.get("STABLE", 0))
    cur_rb = float(cur_grp.get("RB", 0))
    cur_total = float(cur_grp.sum())
    grand["current_res"] = round((cur_stable + cur_rb) / cur_total * 100, 2) if cur_total else 0
    
    return {"bucket": bucket, "teams": teams, "grand_total": grand}


@app.get("/api/ranking")
async def ranking():
    df = load_data()
    if df is None:
        raise HTTPException(status_code=500, detail="Data file not found")
    
    rows = []
    for team, grp in df.groupby("TEAM", dropna=False):
        if not pd.notna(team):
            continue
        total = len(grp)
        paid = int((grp["RECEIPT CUT"] == "PAID").sum())
        stats = resolution_stats(grp)
        collection = float(grp["Paid Amount"].sum())
        rows.append({
            "executive": team,
            "cases": total,
            "paid": paid,
            "unpaid": total - paid,
            "receipt_pct": round(paid / total * 100, 2) if total else 0,
            "resolution_pct": stats["resolution_pct"],
            "collection": round(collection, 2),
        })
    
    rows.sort(key=lambda x: x["resolution_pct"], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    
    return {"ranking": rows}


# ── PAYMENT UPDATE API ──
import time as _time_mod

GSHEET_WEBHOOK = os.environ.get("GSHEET_WEBHOOK", "https://script.google.com/macros/s/AKfycbyKveilFfsklkMv6Q0FpWC-Y2RtYi6jkZWKBwxgeGifIP6L-71XcmWMOaNdZOushRDwag/exec")

_PAYMENT_QUEUE_FILE = Path(os.environ.get("PAYMENT_QUEUE_PATH", "/tmp/payment_queue.json" if CLOUD_MODE else str(APP_DIR / "payment_queue.json")))

def _load_payment_queue():
    try:
        if _PAYMENT_QUEUE_FILE.exists():
            with open(_PAYMENT_QUEUE_FILE, "r", encoding="utf-8") as f:
                return _json.load(f)
    except Exception:
        pass
    return []

def _save_payment_queue(queue):
    try:
        with open(_PAYMENT_QUEUE_FILE, "w", encoding="utf-8") as f:
            _json.dump(queue, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ Payment queue save failed: {e}")

@app.post("/api/payment-update")
async def payment_update(request: Request):
    """Update payment: save to Google Sheets + update in-memory cache."""
    body = await request.json()
    loan_no = str(body.get("loan_no", "")).strip()
    amount = body.get("amount", 0)
    mode = str(body.get("mode", "")).strip().upper()
    pay_date = str(body.get("date", "")).strip()
    
    if not loan_no or not amount or not mode or not pay_date:
        return {"status": "error", "message": "All fields required (loan_no, amount, mode, date)"}
    
    # Validate mode
    valid_modes = ["COLLECT", "CASH", "ONLINE", "CHEQUE"]
    if mode not in valid_modes:
        return {"status": "error", "message": f"Invalid mode. Use: {', '.join(valid_modes)}"}
    
    # 1. Save to Google Sheets (persistent, lightweight)
    try:
        import requests as _req
        gsheet_data = {"loan_no": loan_no, "amount": float(amount), "mode": mode, "date": pay_date}
        _req.post(GSHEET_WEBHOOK, json=gsheet_data, timeout=10)
        print(f"💳 Payment saved to GSheets: {loan_no}")
    except Exception as e:
        print(f"⚠️ GSheet save failed: {e}")
        # Still continue — save to local queue as backup
    
    # 2. Update in-memory DataFrame (report instant update, zero memory spike)
    df = _cache.get("df")
    if df is not None and "LOAN NO" in df.columns:
        mask = df["LOAN NO"].astype(str).str.strip() == loan_no
        if mask.any():
            idx = df[mask].index[0]
            if "Mode Of Payment" in df.columns:
                df.at[idx, "Mode Of Payment"] = mode
            if "Payment Date" in df.columns:
                df.at[idx, "Payment Date"] = pay_date
            if "Paid Amount" in df.columns:
                df.at[idx, "Paid Amount"] = float(amount)
            if "RECEIPT CUT" in df.columns:
                df.at[idx, "RECEIPT CUT"] = "PAID"
            if "POS STATUS" in df.columns:
                df.at[idx, "POS STATUS"] = "STABLE"
            print(f"✅ In-memory updated: {loan_no}")
        else:
            return {"status": "error", "message": f"Loan No '{loan_no}' not found"}
    else:
        return {"status": "error", "message": "Data not loaded yet"}
    
    # 3. Save to local queue (backup + for sync_worker)
    queue = _load_payment_queue()
    queue.append({
        "loan_no": loan_no,
        "amount": float(amount),
        "mode": mode,
        "date": pay_date,
        "timestamp": _dt.utcnow().isoformat(),
        "synced": False
    })
    _save_payment_queue(queue)
    
    return {"status": "ok", "message": f"✅ Payment updated for Loan {loan_no}"}


@app.get("/api/payment-queue")
async def get_payment_queue():
    """Get pending payment queue (not yet synced to HDFC)."""
    queue = _load_payment_queue()
    pending = [q for q in queue if not q.get("synced")]
    return {"pending": len(pending), "entries": pending}


@app.post("/api/payment-queue/clear")
async def clear_synced_payments():
    """Clear synced entries from queue (called by sync_worker after successful HDFC write)."""
    queue = _load_payment_queue()
    pending = [q for q in queue if not q.get("synced")]
    _save_payment_queue(pending)
    return {"status": "ok", "remaining": len(pending)}


@app.post("/api/sync-to-main")
async def sync_to_main():
    """Trigger sync from GSheets to main Excel files (PC must be on, sync_worker running)."""
    trigger_file = APP_DIR / "sync_trigger.txt"
    try:
        trigger_file.write_text(f"triggered at {_dt.utcnow().isoformat()}")
        return {"status": "ok", "message": "✅ Sync triggered! Worker will process in next 30 sec."}
    except Exception as e:
        return {"status": "error", "message": f"❌ Failed: {str(e)}"}


# Root → device-based redirect (with query params forwarding)
@app.get("/")
async def root(request: Request):
    user_agent = request.headers.get("user-agent", "").lower()
    query_string = str(request.query_params)
    suffix = f"?{query_string}" if query_string else ""

    mobile_keywords = [
        "android",
        "iphone",
        "ipad",
        "mobile"
    ]

    if any(keyword in user_agent for keyword in mobile_keywords):
        return RedirectResponse(
            url=f"/mobile/{suffix}",
            status_code=302
        )

    return RedirectResponse(
        url=f"https://dashboard.rccapp.xyz{suffix}",
        status_code=302
    )


if __name__ == "__main__":
    # Pre-load data at startup so first request is instant
    print("⏳ Pre-loading Excel data...")
    load_data()
    port = int(os.environ.get("PORT", 8000))
    print(f"\n🚀 RCC Mobile PWA running!")
    print(f"📱 Open on phone: http://<your-pc-ip>:{port}")
    print(f"💻 Local: http://localhost:{port}\n")
    uvicorn.run(app, host="0.0.0.0", port=port)
